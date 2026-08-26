from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from sqlalchemy import select

from portfolio_agent.bootstrap import Runtime
from portfolio_agent.cbit_contract import (
    CBIT_PROFILE_KEY,
    CBIT_ROWS,
    CbitRowRole,
    CbitValueShape,
)
from portfolio_agent.cli import _parser
from portfolio_agent.enums import DataClassification
from portfolio_agent.models import (
    CompanyProgrammeMembershipModel,
    ObservationModel,
    ObservationNarrativeModel,
)


def _synthetic_value(shape: CbitValueShape) -> str | int:
    values: dict[CbitValueShape, str | int] = {
        CbitValueShape.INTEGER: 2,
        CbitValueShape.CURRENCY: "GBP 100",
        CbitValueShape.PERCENTAGE: "12%",
        CbitValueShape.ORDINAL: 6,
        CbitValueShape.REPORTED_DURATION: "3 synthetic weeks",
        CbitValueShape.TEXT: "Synthetic response",
        CbitValueShape.LIST: "Synthetic item A; Synthetic item B",
    }
    return values[shape]


def _structural_twin() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic CBIT twin"
    sheet.cell(row=1, column=2, value="Alpha Synthetic")
    sheet.cell(row=1, column=3, value="Beta Synthetic")
    for definition in CBIT_ROWS:
        sheet.cell(row=definition.row_number, column=1, value=definition.label)
        for column in (2, 3):
            if definition.role is CbitRowRole.IDENTITY:
                number = "00000001" if column == 2 else "00000002"
                sheet.cell(
                    row=definition.row_number,
                    column=column,
                    value=f"Synthetic Company {column} Ltd {number}",
                )
            elif definition.role is CbitRowRole.JOIN_PERIOD:
                sheet.cell(row=definition.row_number, column=column, value="2024 Q1")
            elif definition.role is CbitRowRole.INPUT:
                sheet.cell(
                    row=definition.row_number,
                    column=column,
                    value=_synthetic_value(definition.value_shape),
                )
            elif definition.role is CbitRowRole.NARRATIVE:
                sheet.cell(
                    row=definition.row_number,
                    column=column,
                    value="Synthetic narrative provenance",
                )
            elif definition.role is CbitRowRole.HELD:
                sheet.cell(
                    row=definition.row_number,
                    column=column,
                    value="10 or 5% synthetic ambiguous value",
                )
            elif definition.role is CbitRowRole.DERIVED:
                sheet.cell(row=definition.row_number, column=column, value="=1+1")
    sheet.cell(row=103, column=1, value="Synthetic unknown extension")
    sheet.cell(row=103, column=2, value="first")
    sheet.cell(row=103, column=3, value="second")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_cbit_structural_twin_imports_inputs_and_holds_ambiguous_rows(
    runtime: Runtime,
) -> None:
    result = runtime.importer.import_bytes(
        _structural_twin(),
        filename="synthetic-cbit-twin.xlsx",
        period_label="SYN-CBIT-2025-Q2",
        reporting_cutoff=date(2025, 6, 30),
        classification=DataClassification.SYNTHETIC,
    )

    input_count = sum(row.role is CbitRowRole.INPUT for row in CBIT_ROWS)
    narrative_count = sum(row.role is CbitRowRole.NARRATIVE for row in CBIT_ROWS)
    held_count = sum(row.role is CbitRowRole.HELD for row in CBIT_ROWS)
    assert result.profile_key == CBIT_PROFILE_KEY
    assert result.company_count == 2
    assert result.observation_count == input_count * 2
    assert result.narrative_count == narrative_count * 2
    assert result.held_field_count == held_count * 2
    assert result.formula_cell_count == 2
    assert result.identity_hold_count == 2
    assert result.programme_start_count == 2

    unknown = [issue for issue in result.issues if issue.code == "unknown_metric"]
    assert len(unknown) == 1
    assert unknown[0].occurrences == 2
    formulas = [issue for issue in result.issues if issue.code == "formula_held"]
    assert len(formulas) == 1
    assert formulas[0].occurrences == 2

    with runtime.session_factory() as session:
        observations = list(
            session.scalars(
                select(ObservationModel).where(
                    ObservationModel.raw_submission_id == result.raw_submission_id
                )
            ).all()
        )
        narratives = list(
            session.scalars(
                select(ObservationNarrativeModel).where(
                    ObservationNarrativeModel.raw_submission_id == result.raw_submission_id
                )
            ).all()
        )
        memberships = list(
            session.scalars(
                select(CompanyProgrammeMembershipModel).where(
                    CompanyProgrammeMembershipModel.raw_submission_id == result.raw_submission_id
                )
            ).all()
        )
    assert all(observation.source_cell not in {"B18", "C18"} for observation in observations)
    assert all(
        not (
            isinstance(observation.original_value_json, str)
            and observation.original_value_json.startswith("=")
        )
        for observation in observations
    )
    assert len(narratives) == narrative_count * 2
    assert {membership.programme_start_date for membership in memberships} == {date(2024, 1, 1)}


def test_cbit_profile_requires_reporting_cutoff(runtime: Runtime) -> None:
    try:
        runtime.importer.import_bytes(
            _structural_twin(),
            filename="synthetic-cbit-twin.xlsx",
            period_label="SYN-CBIT-NO-CUTOFF",
            classification=DataClassification.SYNTHETIC,
        )
    except ValueError as exc:
        assert "reporting cutoff" in str(exc)
    else:
        raise AssertionError("CBIT import must fail closed without a reporting cutoff")


def test_cli_parses_explicit_historical_cutoff_for_cbit_import() -> None:
    args = _parser().parse_args(
        [
            "import",
            "synthetic-cbit-twin.xlsx",
            "--period",
            "SYN-CBIT-2025-Q2",
            "--cutoff",
            "2025-06-30",
        ]
    )
    assert args.cutoff == date(2025, 6, 30)
