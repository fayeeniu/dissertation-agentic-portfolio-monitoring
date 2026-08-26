from __future__ import annotations

import csv
import io
import json
import os
import re
from collections import OrderedDict
from collections.abc import Iterable
from contextlib import suppress
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session, sessionmaker

from .catalogue import MetricCatalogue, seed_catalogue
from .cbit_contract import (
    CBIT_CONTRACT_VERSION,
    CBIT_PROFILE_KEY,
    CBIT_ROWS_BY_LABEL,
    CbitRowRole,
    canonicalize_label,
    detect_cbit_profile,
)
from .enums import DataClassification, IdentifierScheme
from .identity import parse_companies_house_identity, resolve_company_identity
from .ids import dataset_id_for, sha256_bytes
from .models import (
    CompanyProgrammeMembershipModel,
    MetricDefinitionModel,
    ObservationModel,
    ObservationNarrativeModel,
    RawSubmissionModel,
    ReportingPeriodModel,
)
from .normalization import normalize_value
from .schemas import ImportIssue, ImportResult


class ImportValidationError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ParsedMetric:
    label: str
    value: Any
    location: str | None
    row_number: int | None = None
    role: CbitRowRole | None = None
    metric_key: str | None = None
    narrative_for: str | None = None
    is_formula: bool = False


@dataclass(frozen=True, slots=True)
class ParsedCompany:
    name: str
    external_id: str | None
    identifier_scheme: IdentifierScheme
    metrics: tuple[ParsedMetric, ...]
    programme_start_value: Any = None
    programme_start_location: str | None = None


@dataclass(frozen=True, slots=True)
class ParsedSubmission:
    companies: tuple[ParsedCompany, ...]
    profile_key: str | None = None
    profile_version: str | None = None


@dataclass(slots=True)
class _IssueAggregate:
    severity: str
    code: str
    message: str
    location: str | None
    occurrences: int = 1


class _IssueAccumulator:
    def __init__(self) -> None:
        self._items: OrderedDict[tuple[str, str], _IssueAggregate] = OrderedDict()

    def add(
        self,
        *,
        severity: str,
        code: str,
        message: str,
        location: str | None = None,
        aggregate_key: str | None = None,
    ) -> None:
        key = (code, aggregate_key or location or message)
        existing = self._items.get(key)
        if existing is not None:
            existing.occurrences += 1
            return
        self._items[key] = _IssueAggregate(severity, code, message, location)

    def as_tuple(self) -> tuple[ImportIssue, ...]:
        return tuple(
            ImportIssue(
                severity=item.severity,
                code=item.code,
                message=item.message,
                location=item.location,
                occurrences=item.occurrences,
            )
            for item in self._items.values()
        )


def _safe_filename(filename: str) -> str:
    basename = Path(filename).name
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", basename).strip("._")
    return safe or "submission.bin"


def _date_or_none(value: Any) -> date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        raise ImportValidationError("Reporting-period dates must use ISO 8601 strings.")
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ImportValidationError(f"Invalid reporting-period date: {value}") from exc


def _has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


_PROGRAMME_QUARTER_PATTERNS = (
    re.compile(r"^(?P<year>\d{4})\s*[-/]?\s*[Qq](?P<quarter>[1-4])$"),
    re.compile(r"^[Qq](?P<quarter>[1-4])\s*[-/]?\s*(?P<year>\d{4})$"),
)


def _programme_start(value: Any) -> tuple[date | None, str | None]:
    """Parse an explicit programme period to its deterministic first day.

    Quarter-only values identify an interval, so the persisted start is the first
    day of that quarter. Unrecognised non-blank values are held rather than guessed.
    """

    if value is None or (isinstance(value, str) and not value.strip()):
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if not isinstance(value, str):
        return None, "invalid_programme_start"
    text = value.strip()
    try:
        return date.fromisoformat(text), None
    except ValueError:
        pass
    for pattern in _PROGRAMME_QUARTER_PATTERNS:
        match = pattern.fullmatch(text)
        if match is not None:
            quarter = int(match.group("quarter"))
            return date(int(match.group("year")), (quarter - 1) * 3 + 1, 1), None
    return None, "invalid_programme_start"


class PortfolioImporter:
    def __init__(
        self,
        session_factory: sessionmaker[Session],
        raw_data_dir: Path,
        catalogue: MetricCatalogue | None = None,
    ) -> None:
        self._session_factory = session_factory
        self._raw_data_dir = raw_data_dir.resolve()
        self._catalogue = catalogue or MetricCatalogue()

    def import_file(
        self,
        path: Path,
        *,
        period_label: str | None = None,
        reporting_cutoff: date | None = None,
        classification: DataClassification = DataClassification.RESTRICTED,
    ) -> ImportResult:
        payload = path.read_bytes()
        return self.import_bytes(
            payload,
            filename=path.name,
            period_label=period_label,
            reporting_cutoff=reporting_cutoff,
            classification=classification,
        )

    def import_bytes(
        self,
        payload: bytes,
        *,
        filename: str,
        period_label: str | None = None,
        reporting_cutoff: date | None = None,
        classification: DataClassification = DataClassification.RESTRICTED,
    ) -> ImportResult:
        suffix = Path(filename).suffix.lower().lstrip(".")
        if suffix not in {"xlsx", "csv", "json"}:
            raise ImportValidationError("Only XLSX, CSV, and JSON submissions are supported.")

        parsed_period_label = period_label
        start_date: date | None = None
        end_date: date | None = None
        embedded_classification: str | None = None
        if suffix == "json":
            (
                parsed,
                parsed_period_label,
                start_date,
                end_date,
                embedded_classification,
            ) = self._parse_json(payload, period_label)
        else:
            if not parsed_period_label or not parsed_period_label.strip():
                raise ImportValidationError(
                    "A reporting-period label is required for XLSX and CSV."
                )
            parsed = self._parse_xlsx(payload) if suffix == "xlsx" else self._parse_csv(payload)

        if embedded_classification and embedded_classification != classification.value:
            raise ImportValidationError(
                "Requested classification conflicts with the submission classification."
            )
        assert parsed_period_label is not None
        parsed_period_label = parsed_period_label.strip()
        effective_cutoff = reporting_cutoff or end_date
        if parsed.profile_key == CBIT_PROFILE_KEY and effective_cutoff is None:
            raise ImportValidationError(
                "The CBIT workbook profile requires an explicit reporting cutoff date."
            )

        content_hash = sha256_bytes(payload)
        dataset_id = dataset_id_for(payload, parsed_period_label)
        issue_accumulator = _IssueAccumulator()
        if effective_cutoff is None:
            issue_accumulator.add(
                severity="warning",
                code="missing_reporting_cutoff",
                message=(
                    "No reporting cutoff was supplied; public evidence collection will fail closed."
                ),
                aggregate_key="submission",
            )

        created_snapshot: Path | None = None
        try:
            with self._session_factory.begin() as session:
                seed_catalogue(session, self._catalogue)
                session.flush()
                period = self._get_or_create_period(
                    session, parsed_period_label, start_date=start_date, end_date=end_date
                )
                existing = session.scalar(
                    select(RawSubmissionModel).where(
                        RawSubmissionModel.sha256 == content_hash,
                        RawSubmissionModel.reporting_period_id == period.id,
                    )
                )
                if existing is not None:
                    self._assert_reimport_compatible(
                        existing,
                        classification=classification,
                        reporting_cutoff=effective_cutoff,
                        profile_key=parsed.profile_key,
                    )
                    (
                        company_count,
                        observation_count,
                        narrative_count,
                        programme_start_count,
                    ) = self._existing_counts(session, existing.id)
                    summary = existing.import_summary_json
                    return ImportResult(
                        dataset_id=existing.dataset_id,
                        raw_submission_id=existing.id,
                        reporting_period_id=period.id,
                        company_count=company_count,
                        observation_count=observation_count,
                        reused_existing=True,
                        profile_key=existing.profile_key,
                        profile_version=existing.profile_version,
                        reporting_cutoff=existing.reporting_cutoff,
                        narrative_count=narrative_count,
                        held_field_count=int(summary.get("held_field_count", 0)),
                        formula_cell_count=int(summary.get("formula_cell_count", 0)),
                        identity_hold_count=int(summary.get("identity_hold_count", 0)),
                        programme_start_count=programme_start_count,
                    )

                snapshot_path, snapshot_was_created = self._write_snapshot(
                    dataset_id, filename, payload, content_hash
                )
                if snapshot_was_created:
                    created_snapshot = snapshot_path
                raw = RawSubmissionModel(
                    dataset_id=dataset_id,
                    reporting_period_id=period.id,
                    source_format=suffix,
                    original_filename=_safe_filename(filename),
                    sha256=content_hash,
                    snapshot_path=str(snapshot_path),
                    classification=classification.value,
                    reporting_cutoff=effective_cutoff,
                    profile_key=parsed.profile_key,
                    profile_version=parsed.profile_version,
                    catalogue_version=self._catalogue.version,
                    catalogue_sha256=self._catalogue.sha256,
                    import_summary_json={},
                )
                session.add(raw)
                session.flush()

                definitions = {
                    row.key: row for row in session.scalars(select(MetricDefinitionModel)).all()
                }
                seen_observations: set[tuple[str, str]] = set()
                imported_company_ids: set[str] = set()
                observation_count = 0
                narrative_count = 0
                held_field_count = 0
                formula_cell_count = 0
                identity_hold_count = 0
                programme_start_count = 0

                for parsed_company in parsed.companies:
                    resolution = resolve_company_identity(
                        session,
                        raw_submission_id=raw.id,
                        name=parsed_company.name,
                        external_id=parsed_company.external_id,
                        identifier_scheme=parsed_company.identifier_scheme,
                        classification=classification,
                    )
                    if resolution.issue_code is not None:
                        issue_accumulator.add(
                            severity="warning" if resolution.company is not None else "error",
                            code=resolution.issue_code,
                            message=resolution.issue_message or resolution.issue_code,
                            aggregate_key=resolution.issue_code,
                        )
                        identity_hold_count += 1
                    company = resolution.company
                    if company is None:
                        continue
                    imported_company_ids.add(company.id)
                    programme_start, programme_start_issue = _programme_start(
                        parsed_company.programme_start_value
                    )
                    if programme_start_issue is not None:
                        issue_accumulator.add(
                            severity="warning",
                            code=programme_start_issue,
                            message=(
                                "Programme start was not an ISO date or an unambiguous quarter "
                                "label; cumulative metrics will abstain from public support."
                            ),
                            location=parsed_company.programme_start_location,
                            aggregate_key=parsed_company.programme_start_location
                            or parsed_company.name,
                        )
                    elif programme_start is not None:
                        if effective_cutoff is not None and programme_start > effective_cutoff:
                            issue_accumulator.add(
                                severity="warning",
                                code="programme_start_after_cutoff",
                                message=(
                                    "Programme start follows the reporting cutoff; cumulative "
                                    "metrics will abstain from public support."
                                ),
                                location=parsed_company.programme_start_location,
                                aggregate_key=parsed_company.programme_start_location
                                or parsed_company.name,
                            )
                        else:
                            session.add(
                                CompanyProgrammeMembershipModel(
                                    raw_submission_id=raw.id,
                                    company_id=company.id,
                                    programme_start_date=programme_start,
                                    submitted_period_label=str(
                                        parsed_company.programme_start_value
                                    ).strip(),
                                    source_cell=parsed_company.programme_start_location,
                                )
                            )
                            programme_start_count += 1
                    observations_by_key: dict[str, ObservationModel] = {}
                    narratives: list[ParsedMetric] = []

                    for parsed_metric in parsed_company.metrics:
                        if parsed_metric.role is CbitRowRole.NARRATIVE:
                            if _has_value(parsed_metric.value):
                                narratives.append(parsed_metric)
                            continue
                        if parsed_metric.is_formula or parsed_metric.role is CbitRowRole.DERIVED:
                            if _has_value(parsed_metric.value):
                                formula_cell_count += 1
                                issue_accumulator.add(
                                    severity="info",
                                    code="formula_held",
                                    message=(
                                        "Derived/formula row held outside canonical facts: "
                                        f"{parsed_metric.label}"
                                    ),
                                    location=(
                                        f"row {parsed_metric.row_number}"
                                        if parsed_metric.row_number
                                        else parsed_metric.location
                                    ),
                                    aggregate_key=parsed_metric.label,
                                )
                            continue
                        if parsed_metric.role is CbitRowRole.HELD:
                            if _has_value(parsed_metric.value):
                                held_field_count += 1
                                issue_accumulator.add(
                                    severity="warning",
                                    code="mixed_field_held",
                                    message=(
                                        "Ambiguous mixed-shape field requires split collection: "
                                        f"{parsed_metric.label}"
                                    ),
                                    location=(
                                        f"row {parsed_metric.row_number}"
                                        if parsed_metric.row_number
                                        else parsed_metric.location
                                    ),
                                    aggregate_key=parsed_metric.label,
                                )
                            continue

                        metric = (
                            self._catalogue.get(parsed_metric.metric_key)
                            if parsed_metric.metric_key
                            else self._catalogue.resolve(parsed_metric.label)
                        )
                        if metric is None:
                            if _has_value(parsed_metric.value):
                                issue_accumulator.add(
                                    severity="warning",
                                    code="unknown_metric",
                                    message=f"Unmapped metric label: {parsed_metric.label}",
                                    location=(
                                        f"row {parsed_metric.row_number}"
                                        if parsed_metric.row_number
                                        else parsed_metric.location
                                    ),
                                    aggregate_key=parsed_metric.label,
                                )
                            continue
                        key = (company.id, metric.key)
                        if key in seen_observations:
                            issue_accumulator.add(
                                severity="error",
                                code="duplicate_observation",
                                message=(
                                    f"Duplicate value for {metric.key}; later value was not "
                                    "imported."
                                ),
                                location=parsed_metric.location,
                                aggregate_key=f"{company.id}:{metric.key}",
                            )
                            continue
                        seen_observations.add(key)
                        normalized = normalize_value(parsed_metric.value, metric)
                        definition = definitions[metric.key]
                        observation = ObservationModel(
                            company_id=company.id,
                            metric_definition_id=definition.id,
                            reporting_period_id=period.id,
                            raw_submission_id=raw.id,
                            original_value_json=parsed_metric.value,
                            normalized_value_json=normalized.value,
                            missing_state=normalized.missing_state.value,
                            unit=normalized.unit,
                            currency=normalized.currency,
                            source_cell=parsed_metric.location,
                            normalization_issue_code=normalized.issue_code,
                            normalization_issue_message=normalized.issue_message,
                        )
                        session.add(observation)
                        session.flush()
                        observations_by_key[metric.key] = observation
                        observation_count += 1

                    for narrative in narratives:
                        assert narrative.narrative_for is not None
                        parent = observations_by_key.get(narrative.narrative_for)
                        session.add(
                            ObservationNarrativeModel(
                                observation_id=parent.id if parent else None,
                                raw_submission_id=raw.id,
                                company_id=company.id,
                                parent_metric_key=narrative.narrative_for,
                                body=str(narrative.value),
                                source_label=narrative.label,
                                source_cell=narrative.location or "unknown",
                            )
                        )
                        narrative_count += 1

                issues = issue_accumulator.as_tuple()
                raw.import_summary_json = {
                    "company_count": len(imported_company_ids),
                    "observation_count": observation_count,
                    "narrative_count": narrative_count,
                    "held_field_count": held_field_count,
                    "formula_cell_count": formula_cell_count,
                    "identity_hold_count": identity_hold_count,
                    "programme_start_count": programme_start_count,
                    "issue_codes": sorted({issue.code for issue in issues}),
                }
                return ImportResult(
                    dataset_id=dataset_id,
                    raw_submission_id=raw.id,
                    reporting_period_id=period.id,
                    company_count=len(imported_company_ids),
                    observation_count=observation_count,
                    issues=issues,
                    profile_key=parsed.profile_key,
                    profile_version=parsed.profile_version,
                    reporting_cutoff=effective_cutoff,
                    narrative_count=narrative_count,
                    held_field_count=held_field_count,
                    formula_cell_count=formula_cell_count,
                    identity_hold_count=identity_hold_count,
                    programme_start_count=programme_start_count,
                )
        except Exception:
            if created_snapshot is not None:
                self._remove_failed_snapshot(created_snapshot, content_hash)
            raise

    @staticmethod
    def _assert_reimport_compatible(
        existing: RawSubmissionModel,
        *,
        classification: DataClassification,
        reporting_cutoff: date | None,
        profile_key: str | None,
    ) -> None:
        if existing.classification != classification.value:
            raise ImportValidationError(
                "Immutable submission already exists under a different classification."
            )
        if reporting_cutoff is not None and existing.reporting_cutoff != reporting_cutoff:
            raise ImportValidationError(
                "Immutable submission already exists under a different reporting cutoff."
            )
        if existing.profile_key != profile_key:
            raise ImportValidationError(
                "Immutable submission profile differs from the stored import profile."
            )

    def _write_snapshot(
        self, dataset_id: str, filename: str, payload: bytes, expected_hash: str
    ) -> tuple[Path, bool]:
        target_dir = self._raw_data_dir / dataset_id
        target_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        target = target_dir / _safe_filename(filename)
        if target.exists():
            if sha256_bytes(target.read_bytes()) != expected_hash:
                raise ImportValidationError(
                    "Immutable snapshot path already contains different bytes."
                )
            return target, False
        with target.open("xb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        target.chmod(0o600)
        return target, True

    @staticmethod
    def _remove_failed_snapshot(path: Path, expected_hash: str) -> None:
        if path.is_file() and sha256_bytes(path.read_bytes()) == expected_hash:
            path.unlink()
            with suppress(OSError):
                path.parent.rmdir()

    @staticmethod
    def _get_or_create_period(
        session: Session,
        label: str,
        *,
        start_date: date | None,
        end_date: date | None,
    ) -> ReportingPeriodModel:
        period = session.scalar(
            select(ReportingPeriodModel).where(ReportingPeriodModel.label == label)
        )
        if period is not None:
            if (start_date and period.start_date and start_date != period.start_date) or (
                end_date and period.end_date and end_date != period.end_date
            ):
                raise ImportValidationError("Reporting-period label conflicts with stored dates.")
            return period
        if start_date and end_date and start_date > end_date:
            raise ImportValidationError("Reporting-period start date must not follow its end date.")
        period = ReportingPeriodModel(label=label, start_date=start_date, end_date=end_date)
        session.add(period)
        session.flush()
        return period

    @staticmethod
    def _existing_counts(session: Session, raw_submission_id: str) -> tuple[int, int, int, int]:
        company_count = session.scalar(
            select(func.count(func.distinct(ObservationModel.company_id))).where(
                ObservationModel.raw_submission_id == raw_submission_id
            )
        )
        observation_count = session.scalar(
            select(func.count(ObservationModel.id)).where(
                ObservationModel.raw_submission_id == raw_submission_id
            )
        )
        narrative_count = session.scalar(
            select(func.count(ObservationNarrativeModel.id)).where(
                ObservationNarrativeModel.raw_submission_id == raw_submission_id
            )
        )
        programme_start_count = session.scalar(
            select(func.count(CompanyProgrammeMembershipModel.id)).where(
                CompanyProgrammeMembershipModel.raw_submission_id == raw_submission_id
            )
        )
        return (
            int(company_count or 0),
            int(observation_count or 0),
            int(narrative_count or 0),
            int(programme_start_count or 0),
        )

    @staticmethod
    def _parse_json(
        payload: bytes, override_period: str | None
    ) -> tuple[ParsedSubmission, str, date | None, date | None, str | None]:
        try:
            document = json.loads(payload)
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ImportValidationError("Submission is not valid UTF-8 JSON.") from exc
        if not isinstance(document, dict):
            raise ImportValidationError("JSON submission must be an object.")
        period = document.get("reporting_period")
        if not isinstance(period, dict):
            period = {}
        embedded_label = period.get("label")
        label = override_period or embedded_label
        if not isinstance(label, str) or not label.strip():
            raise ImportValidationError("JSON submission requires reporting_period.label.")
        if override_period and embedded_label and override_period != embedded_label:
            raise ImportValidationError("Requested period conflicts with embedded JSON period.")
        embedded_classification = document.get("classification")
        if embedded_classification is not None and not isinstance(embedded_classification, str):
            raise ImportValidationError("JSON classification must be a string when present.")
        companies = document.get("companies")
        if not isinstance(companies, list) or not companies:
            raise ImportValidationError("JSON submission requires a non-empty companies array.")
        parsed: list[ParsedCompany] = []
        for index, item in enumerate(companies):
            if not isinstance(item, dict) or not isinstance(item.get("name"), str):
                raise ImportValidationError(f"companies[{index}] requires a string name.")
            metrics = item.get("metrics")
            if not isinstance(metrics, dict):
                raise ImportValidationError(f"companies[{index}].metrics must be an object.")
            external_id = item.get("external_id")
            if external_id is not None and not isinstance(external_id, str):
                raise ImportValidationError(f"companies[{index}].external_id must be a string.")
            programme_start_value = item.get("programme_start_date")
            if programme_start_value is not None and not isinstance(programme_start_value, str):
                raise ImportValidationError(
                    f"companies[{index}].programme_start_date must be a string."
                )
            parsed.append(
                ParsedCompany(
                    name=item["name"],
                    external_id=external_id,
                    identifier_scheme=IdentifierScheme.LEGACY,
                    metrics=tuple(
                        ParsedMetric(
                            label=str(key), value=value, location=f"companies[{index}].{key}"
                        )
                        for key, value in metrics.items()
                    ),
                    programme_start_value=programme_start_value,
                    programme_start_location=f"companies[{index}].programme_start_date",
                )
            )
        return (
            ParsedSubmission(tuple(parsed)),
            label,
            _date_or_none(period.get("start_date")),
            _date_or_none(period.get("end_date")),
            embedded_classification,
        )

    @staticmethod
    def _parse_xlsx(payload: bytes) -> ParsedSubmission:
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        except Exception as exc:
            raise ImportValidationError("Workbook could not be opened as XLSX.") from exc
        try:
            worksheet = workbook.active
            rows = [list(row) for row in worksheet.iter_rows()]
            labels_by_row = {
                index: str(row[0].value).strip()
                for index, row in enumerate(rows, start=1)
                if row and row[0].value is not None and str(row[0].value).strip()
            }
            if detect_cbit_profile(labels_by_row):
                return PortfolioImporter._parse_cbit_rows(rows)
            return ParsedSubmission(
                PortfolioImporter._parse_cell_matrix_rows(rows, location_style="xlsx")
            )
        finally:
            workbook.close()

    @staticmethod
    def _parse_cbit_rows(rows: list[list[Cell]]) -> ParsedSubmission:
        if not rows or len(rows[0]) < 2:
            raise ImportValidationError("CBIT workbook has no company columns.")
        company_columns: list[tuple[int, str]] = []
        for column_index, cell in enumerate(rows[0][1:], start=2):
            if cell.value is None or not str(cell.value).strip():
                continue
            company_columns.append((column_index, str(cell.value).strip()))
        if not company_columns:
            raise ImportValidationError("CBIT workbook has no company columns.")

        parsed: list[ParsedCompany] = []
        for column_index, header_name in company_columns:
            identity_value = (
                rows[2][column_index - 1].value
                if len(rows) >= 3 and len(rows[2]) >= column_index
                else None
            )
            official_name, company_number = parse_companies_house_identity(
                identity_value, fallback_name=header_name
            )
            metrics: list[ParsedMetric] = []
            programme_start_value: Any = None
            programme_start_location: str | None = None
            for row_number, row_cells in enumerate(rows[1:], start=2):
                if not row_cells or row_cells[0].value is None:
                    continue
                label = str(row_cells[0].value).strip()
                if not label:
                    continue
                value_cell = row_cells[column_index - 1] if len(row_cells) >= column_index else None
                value = value_cell.value if value_cell is not None else None
                is_formula = bool(
                    value_cell is not None
                    and (
                        value_cell.data_type == "f"
                        or (isinstance(value, str) and value.startswith("="))
                    )
                )
                definition = CBIT_ROWS_BY_LABEL.get(canonicalize_label(label))
                if definition is not None and definition.role is CbitRowRole.JOIN_PERIOD:
                    programme_start_value = value
                    programme_start_location = f"{get_column_letter(column_index)}{row_number}"
                    continue
                if definition is not None and definition.role in {
                    CbitRowRole.SECTION,
                    CbitRowRole.IDENTITY,
                }:
                    continue
                metrics.append(
                    ParsedMetric(
                        label=label,
                        value=value,
                        location=f"{get_column_letter(column_index)}{row_number}",
                        row_number=row_number,
                        role=definition.role if definition else None,
                        metric_key=definition.metric_key if definition else None,
                        narrative_for=definition.narrative_for if definition else None,
                        is_formula=is_formula,
                    )
                )
            parsed.append(
                ParsedCompany(
                    name=official_name,
                    external_id=company_number,
                    identifier_scheme=IdentifierScheme.COMPANIES_HOUSE_NUMBER,
                    metrics=tuple(metrics),
                    programme_start_value=programme_start_value,
                    programme_start_location=programme_start_location,
                )
            )
        return ParsedSubmission(
            companies=tuple(parsed),
            profile_key=CBIT_PROFILE_KEY,
            profile_version=CBIT_CONTRACT_VERSION,
        )

    @staticmethod
    def _parse_csv(payload: bytes) -> ParsedSubmission:
        try:
            decoded = payload.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ImportValidationError("CSV submission must use UTF-8 encoding.") from exc
        rows = list(csv.reader(io.StringIO(decoded)))
        return ParsedSubmission(PortfolioImporter._parse_matrix_rows(rows, location_style="csv"))

    @staticmethod
    def _parse_cell_matrix_rows(
        rows: list[list[Cell]], *, location_style: str
    ) -> tuple[ParsedCompany, ...]:
        values = [[cell.value for cell in row] for row in rows]
        parsed = PortfolioImporter._parse_matrix_rows(values, location_style=location_style)
        formula_locations = {
            cell.coordinate for row in rows for cell in row if cell.data_type == "f"
        }
        return tuple(
            ParsedCompany(
                name=company.name,
                external_id=company.external_id,
                identifier_scheme=company.identifier_scheme,
                metrics=tuple(
                    ParsedMetric(
                        label=metric.label,
                        value=metric.value,
                        location=metric.location,
                        is_formula=metric.location in formula_locations,
                    )
                    for metric in company.metrics
                ),
                programme_start_value=company.programme_start_value,
                programme_start_location=company.programme_start_location,
            )
            for company in parsed
        )

    @staticmethod
    def _parse_matrix_rows(
        rows: Iterable[Iterable[Any]], *, location_style: str
    ) -> tuple[ParsedCompany, ...]:
        materialized = [list(row) for row in rows]
        if not materialized or len(materialized[0]) < 2:
            raise ImportValidationError(
                "Matrix submission requires metric labels in column A and companies across row 1."
            )
        headers = materialized[0]
        companies: list[tuple[int, str]] = []
        for column_index, header in enumerate(headers[1:], start=2):
            if header is None or not str(header).strip():
                continue
            companies.append((column_index, str(header).strip()))
        if not companies:
            raise ImportValidationError("Matrix submission has no company columns.")
        parsed: list[ParsedCompany] = []
        for column_index, company_name in companies:
            metrics: list[ParsedMetric] = []
            for row_index, row in enumerate(materialized[1:], start=2):
                if not row or row[0] is None or not str(row[0]).strip():
                    continue
                value = row[column_index - 1] if column_index - 1 < len(row) else None
                location = (
                    f"{get_column_letter(column_index)}{row_index}"
                    if location_style == "xlsx"
                    else f"row {row_index}, column {column_index}"
                )
                metrics.append(
                    ParsedMetric(label=str(row[0]).strip(), value=value, location=location)
                )
            parsed.append(
                ParsedCompany(
                    name=company_name,
                    external_id=None,
                    identifier_scheme=IdentifierScheme.LEGACY,
                    metrics=tuple(metrics),
                )
            )
        return tuple(parsed)
